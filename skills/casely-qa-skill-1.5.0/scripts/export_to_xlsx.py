"""
Casely Export Module — converts Markdown test case tables into individual Excel files.
"""

import re
import sys
import argparse
from pathlib import Path
from typing import Optional, List, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 60


def _split_table_row(line: str) -> List[str]:
    """Split a Markdown table row into individual cell values."""
    raw_parts = re.split(r'(?<!\\)\|', line)
    start: int = 1 if line.startswith('|') else 0
    end: int = -1 if line.endswith('|') else len(raw_parts)

    result: List[str] = []
    raw_len = len(raw_parts)
    for i in range(start, int(end) if end >= 0 else raw_len + int(end)):
        cell = raw_parts[i]
        if cell is not None:
            result.append(cell.strip().replace(r'\|', '|'))
    return result


def parse_md_table(md_content: str) -> Tuple[List[str], List[List[str]]]:
    """Parse a Markdown table into headers and data rows."""
    lines: List[str] = [
        line.strip()
        for line in md_content.strip().split('\n')
        if line.strip().startswith('|')
    ]
    if len(lines) < 2:
        return [], []

    headers: List[str] = _split_table_row(lines[0])
    rows: List[List[str]] = []

    for line in lines[1:]:
        # Skip separator rows like |---|---|
        if re.match(r'^\|[\s\-:|]+\|$', line):
            continue

        row: List[str] = _split_table_row(line)
        if not row:
            continue

        # Pad short rows with empty strings
        while len(row) < len(headers):
            row.append('')

        # Trim extra columns
        rows.append(row[0:len(headers)])

    return headers, rows


def export_to_xlsx(results_dir: str, output_path: str) -> None:
    """Convert each Markdown test case to a separate Excel file."""
    results_path = Path(results_dir)
    out_dir = Path(output_path)

    if not results_path.exists():
        print(f"Error: Results directory not found: {results_dir}")
        sys.exit(1)

    # Create the export directory if it does not exist
    out_dir.mkdir(parents=True, exist_ok=True)

    md_files = sorted(results_path.glob('*.md'))
    if not md_files:
        print(f"Warning: No .md files found in {results_dir}")
        return

    for md_file in md_files:
        headers, rows = parse_md_table(md_file.read_text(encoding='utf-8'))
        if not headers:
            print(f"Skipping {md_file.name}: No table found.")
            continue

        # Create a new workbook for each file
        wb = Workbook()
        ws_opt = wb.active
        if ws_opt is None:
            print(f"Skipping {md_file.name}: Failed to create worksheet.")
            continue
        ws: Worksheet = ws_opt  # type: ignore[assignment]
        ws.title = "Test Case"

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                clean_value = value.replace('<br>', '\n').replace('<BR>', '\n') if value else ''
                cell = ws.cell(row=row_idx, column=col_idx, value=clean_value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Style headers: bold + center
        header_font = Font(bold=True)
        for col_idx in range(1, len(headers) + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.font = header_font
            header_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-fit column widths
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(ws.cell(row=1, column=col_idx).value or ''))
            total_rows = len(rows)
            for row_idx in range(2, total_rows + 2):
                cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
                # For multiline cells, use the longest line
                for line in cell_val.split('\n'):
                    max_len = max(max_len, len(line))
            col_letter = get_column_letter(col_idx)
            width = min(max(max_len + 2, MIN_COL_WIDTH), MAX_COL_WIDTH)
            ws.column_dimensions[col_letter].width = width

        # Save with the same name but .xlsx extension
        dest_file = out_dir / f"{md_file.stem}.xlsx"
        wb.save(str(dest_file))
        print(f"Exported: {dest_file.name}")



def find_latest_project() -> Optional[Path]:
    """Find the most recently modified project directory."""
    projects_dir = Path("projects")
    if not projects_dir.exists():
        return None

    subdirs = [d for d in projects_dir.iterdir() if d.is_dir()]
    if not subdirs:
        return None

    return max(subdirs, key=lambda d: d.stat().st_mtime)


def main() -> None:
    """CLI interface for the exporter."""
    arg_parser = argparse.ArgumentParser(
        description='Casely Export — Markdown to Excel converter'
    )
    arg_parser.add_argument('results_dir', nargs='?', help='Path to results MD files')
    arg_parser.add_argument('output_path', nargs='?', help='Path to export XLSX files')
    args = arg_parser.parse_args()

    results_dir: Optional[str] = args.results_dir
    output_path: Optional[str] = args.output_path

    # If arguments are not provided, try to find the project automatically
    if not results_dir or not output_path:
        latest = find_latest_project()
        if latest is not None:
            results_dir = str(latest / "results")
            output_path = str(latest / "exports")
            print(f"Auto-detected project: {latest.name}")
        else:
            print("Error: No paths provided and no projects found in 'projects/' directory.")
            sys.exit(1)

    export_to_xlsx(results_dir, output_path)


if __name__ == '__main__':
    main()
